import streamlit as st
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
import os


excel_path = os.path.join('liegenschaftszinssatz_tabelle.xlsx')
wb = load_workbook(excel_path)
ws = wb.active


vpi = {
    '2021': 104.3,
    '2022': 113.5,
    '2023': 117.8,
    '2024': 120.2
}


def show_ertragswertverfahren_rechner(switch_page):
    st.title('Ertragswertverfahren')

    mieten = []
    flaechen = []
    pruefung_mieten = []
    abweichung_vergleichsmiete = []
    neue_mieten = []


    def berechne_bodenwert(bodenrichtwert, grundstuecksgroesse):
        brw = Decimal(str(bodenrichtwert))
        gr = Decimal(str(grundstuecksgroesse))
        bodenwert = brw * gr
        return bodenwert

    def berechne_rohertrag():
        rohertrag = sum(mieten) * 12
        return rohertrag

    def mieten_pruefen(eingabe_miete, eingabe_flaeche, vergleichsmiete):
        miete = Decimal(str(eingabe_miete))
        flaeche = Decimal(str(eingabe_flaeche))

        miete_qm = miete / flaeche

        vgl_miete = Decimal(str(vergleichsmiete))

        verhaeltnis = ((miete_qm - vgl_miete) / vgl_miete)
        betrag_verhaeltnis = abs(verhaeltnis)

        abweichung_vergleichsmiete.append(betrag_verhaeltnis)

        miete_neu = vgl_miete * flaeche
        neue_mieten.append(miete_neu)

        if betrag_verhaeltnis > Decimal('0.2') and not miete == 0:
            miete = miete_neu
            pruefung_mieten.append('v')
        elif miete == 0:
            miete = miete_neu
            pruefung_mieten.append('l')
        else:
            pruefung_mieten.append('ok')

        mieten.append(miete)
        flaechen.append(flaeche)

    def berechne_angepasster_vpi(vpi_aktuell):
        vpi_angepasst = Decimal(str(vpi_aktuell)) / Decimal('77.1')
        return vpi_angepasst.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

    def berechne_verwaltungskosten(vpi_angepasst, anzahl_wohnungen):
        verwaltungskosten = Decimal('230') * vpi_angepasst * anzahl_wohnungen
        return verwaltungskosten.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

    def berechne_instandhaltungskosten(vpi_angepasst):
        flaeche_gesamt = sum(flaechen)
        instandhaltungskosten = Decimal('9') * vpi_angepasst * flaeche_gesamt
        return instandhaltungskosten.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

    def berechne_mietausfallwagnis(rohertrag):
        mietausfallwagnis = Decimal('0.02') * rohertrag
        return mietausfallwagnis.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

    def bestimme_restnutzungsdauer(jahr_bewertungsstichtag, jahr_bezugsfertigkeit):
        alter_gebaeude = Decimal(str(jahr_bewertungsstichtag)) - Decimal(str(jahr_bezugsfertigkeit))
        gesamtnutzungsdauer = 80
        restnutzungsdauer = gesamtnutzungsdauer - alter_gebaeude
        mindestnutzungsdauer = gesamtnutzungsdauer * Decimal('0.3')
        if restnutzungsdauer < mindestnutzungsdauer:
            restnutzungsdauer = mindestnutzungsdauer
        return restnutzungsdauer, mindestnutzungsdauer

    def bestimme_vervielfaeltiger(restnutzungsdauer):
        zielzeile = int(restnutzungsdauer + 5)
        zielzeile = f'G{zielzeile}'
        vervielfaeltiger = ws[zielzeile].value
        return Decimal(vervielfaeltiger).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)


    # Eingaben

    bodenrichtwert = st.number_input('Bodenrichtwert:', min_value=0)

    grundstuecksgroesse = st.number_input('Grundstücksgröße:', min_value=0)

    vergleichsmiete = st.number_input('Vergleichsmiete pro m²:', min_value=1.0, format='%.10f')

    options = ['2021', '2022', '2023', '2024']
    jahr_bewertungsstichtag = st.selectbox('Jahr des Bewertungsstichtages:', options)

    jahr_bezugsfertigkeit = st.number_input('Jahr der Bezugsfertigkeit:', step=1)

    anzahl_wohnungen = st.number_input('Anzahl Wohnungen:', min_value=1, step=1)

    for i in range(anzahl_wohnungen):
        st.write(f'Wohnung {i+1}')
        eingabe_miete = st.number_input(f'Miete:', key=f'a{i}')
        eingabe_flaeche = st.number_input(f'Fläche:', min_value=1, key=f'b{i}')
        mieten_pruefen(eingabe_miete, eingabe_flaeche, vergleichsmiete)

    # Berechnungen

    bodenwert = berechne_bodenwert(bodenrichtwert, grundstuecksgroesse)

    rohertrag = berechne_rohertrag()

    vpi_aktuell = vpi[jahr_bewertungsstichtag]
    vpi_angepasst = berechne_angepasster_vpi(vpi_aktuell)

    verwaltungskosten = berechne_verwaltungskosten(vpi_angepasst, anzahl_wohnungen)

    instandhaltungskosten = berechne_instandhaltungskosten(vpi_angepasst)

    mietausfallwagnis = berechne_mietausfallwagnis(rohertrag)

    bewirtschaftungskosten = verwaltungskosten + instandhaltungskosten + mietausfallwagnis

    reinertrag = rohertrag - bewirtschaftungskosten

    bodenwertverzinsung = Decimal('0.035') * bodenwert

    gebaeudereinertrag = reinertrag - bodenwertverzinsung

    restnutzungsdauer,_ = bestimme_restnutzungsdauer(jahr_bewertungsstichtag, jahr_bezugsfertigkeit)

    verfielfaeltiger = bestimme_vervielfaeltiger(restnutzungsdauer)

    gebaeudeertragswert = verfielfaeltiger * gebaeudereinertrag

    grundstuecksertragswert = bodenwert + gebaeudeertragswert


    # Ergebnisanzeige

    if st.button('Berechnen'):

        st.write('Bodenwert:', f'{bodenwert} €')
        st.write('Rohertrag:', f'{rohertrag:.2f} €')
        st.write('Verwaltungskosten:', f'{verwaltungskosten} €')
        st.write('Instandhaltungskosten:', f'{instandhaltungskosten} €')
        st.write('Mietausfallwagnis:', f'{mietausfallwagnis} €')
        st.write(f'Bewirtschaftungskosten: {bewirtschaftungskosten} €')
        st.write(f'Reinertrag: {reinertrag:.2f} €')
        st.write(f'Bodenwertverzinsung: {bodenwertverzinsung:.2f} € (Liegschaftszinssatz = 3,5%)')
        st.write(f'Gebäudereinertrag: {gebaeudereinertrag:.2f} €')
        st.write(f'Gebäudeertragswert: {gebaeudeertragswert:.2f} €')
        st.write(f'GRUNDSTÜCKSERTRAGSWERT: {grundstuecksertragswert:.2f} €')
        st.write(' ')
        st.write(' ')
        st.write(f'VPI in {jahr_bewertungsstichtag}: {vpi_aktuell}')
        st.write(f'angepasster VPI: {vpi_angepasst}')
        st.write(f'Restnutzungsdauer: {restnutzungsdauer} Jahre')
        st.write(f'Verfielfältiger: {verfielfaeltiger} ')


        nummer = 1
        for i in pruefung_mieten:
            if i == 'v':
                st.info(f'Wohnung {nummer}: Die Miete weicht um mehr als 20% von der Vergleichsmiete ab ({(abweichung_vergleichsmiete[nummer-1]*100):.2f}%), die Vergleichsmiete wird angesetzt')
                nummer += 1
            elif i == 'l':
                st.info(f'Wohnung {nummer}: Leerstand, die Vergleichsmiete wird angesetzt. Die Vergleichsmiete beträgt {neue_mieten[nummer-1]:.2f} €')
                nummer += 1
            elif i == 'ok':
                nummer += 1

        restnutzungsdauer, mindestnutzungsdauer = bestimme_restnutzungsdauer(jahr_bewertungsstichtag, jahr_bezugsfertigkeit)
        if restnutzungsdauer == mindestnutzungsdauer:
            st.warning('Die wirtschaftliche Gesamtnutzungsdauer beträgt 80 Jahre. '
                       'Die Mindestnutzungsdauer beträgt davon 30%, also 24 Jahre. '
                       'Weil die Restnutzungsdauer die Mindestnutzungsdauer unterschreiten würde, '
                       'wird die Restnutzungsdauer auf 24 Jahre festgelegt.')



    if st.button('zurück'):
        switch_page('menu')
